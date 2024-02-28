import boto3
import openpyxl
import base64
import subprocess

aws_region = 'us-east-1'
ec2_client = boto3.client('ec2', region_name=aws_region)
vpc_id = 'vpc-086d917d78186bd1f'

# Step 1: Set Up MongoDB
mongo_url = 'mongodb+srv://surendergupta:ceOC9UhUh5trGjEE@taskmanagementcluster.tsyynim.mongodb.net/travelMemory'

instance_type = 't2.micro'
ami_id = 'ami-0c7217cdde317cfec'
key_pair_name = 'suri-tm-pj'
security_group_name = 'launch-wizard-1'
security_group_id = 'sg-00bc6a6d3f2de3870'
user_data_script = """#!/bin/bash
sudo apt-get update -y
sudo apt-get install -y ca-certificates curl gnupg
sudo mkdir -p /etc/apt/keyrings
curl -fsSL https://deb.nodesource.com/gpgkey/nodesource-repo.gpg.key | sudo gpg --dearmor -o /etc/apt/keyrings/nodesource.gpg
NODE_MAJOR=18
echo "deb [signed-by=/etc/apt/keyrings/nodesource.gpg] https://deb.nodesource.com/node_$NODE_MAJOR.x nodistro main" | sudo tee /etc/apt/sources.list.d/nodesource.list
sudo apt-get update
sudo apt install nodejs -y
cd /home/ubuntu/
sudo apt install git -y
sudo git clone https://github.com/surendergupta/TravelMemory.git
cd /home/ubuntu/TravelMemory/backend/
echo "MONGO_URI='{}'" > .env
echo "PORT=5000" >> .env
npm install
node index.js
""".format(mongo_url)
user_data_encoded = base64.b64encode(user_data_script.encode()).decode('utf-8')

response = ec2_client.run_instances(
    ImageId=ami_id,
    InstanceType=instance_type,
    KeyName=key_pair_name,
    SecurityGroups=[security_group_name],
    UserData=user_data_encoded,
    MinCount=1,
    MaxCount=1,
    TagSpecifications=[
        {
            'ResourceType': 'instance',
            'Tags': [
                {
                    'Key': 'Name',
                    'Value': 'suri-tm-be'
                }
            ]
        }
    ]
)

backend_instance_id = response['Instances'][0]['InstanceId']
backend_instance_type = response['Instances'][0]['InstanceType']
backend_instance_privateip = response['Instances'][0]['PrivateIpAddress']
backend_instance_launchtime = response['Instances'][0]['LaunchTime']
backend_instance_group_name = response['Instances'][0]['SecurityGroups'][0]['GroupName']
backend_instance_group_id = response['Instances'][0]['SecurityGroups'][0]['GroupId']
print(f"Launched backend instance with ID: {backend_instance_id}")

ec2_resource = boto3.resource('ec2')

instance = ec2_resource.Instance(backend_instance_id)
instance.wait_until_running()



try:
    workbook = openpyxl.load_workbook("instance_info_backend.xlsx")
    worksheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Instance Info"
    worksheet['A1'] = "Instance ID"
    worksheet['B1'] = "Instance Type"
    worksheet['C1'] = "Public IPv4 address"
    worksheet['D1'] = "Private IP"
    worksheet['E1'] = "Launch time"
    worksheet['F1'] = "Security group name"

# Add instance information to the worksheet
next_row = worksheet.max_row + 1
worksheet.cell(row=next_row, column=1, value=instance.instance_id)
worksheet.cell(row=next_row, column=2, value=instance.instance_type)
worksheet.cell(row=next_row, column=3, value=instance.public_ip_address)
worksheet.cell(row=next_row, column=4, value=instance.private_ip_address)
worksheet.cell(row=next_row, column=5, value=str(instance.launch_time))
worksheet.cell(row=next_row, column=6, value=instance.security_groups[0]['GroupName'] if instance.security_groups else '')

# Save the Excel file
workbook.save("instance_info_be.xlsx")


# instance = response[0]
# instance.wait_until_running()

subprocess.run(['python', 'boto3_fe.py'])