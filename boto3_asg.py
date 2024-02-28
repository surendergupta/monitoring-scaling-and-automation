import boto3
import base64
import openpyxl
from datetime import datetime
from botocore.exceptions import NoCredentialsError
import subprocess

aws_region = 'us-east-1'
ec2_client = boto3.client('ec2', region_name=aws_region)
elbv2_client = boto3.client('elbv2', region_name=aws_region)
asg_client = boto3.client('autoscaling', region_name=aws_region)
# Auto Scaling Group (ASG) configuration
launch_template_name = 'suri-tm-fe-launch-template'
asg_name = 'suri-tm-asg'
min_size = 1
max_size = 3
desired_capacity = 1

# Scaling policies configuration
cpu_utilization_threshold = 50  # Scale out if CPU utilization is above this threshold
network_traffic_threshold = 5000000  # Scale out if network traffic is above this threshold
scale_out_cooldown = 300  # seconds
scale_in_cooldown = 300  # seconds
target_group_arn1 = None
def get_instances_running_by_name():
    instance_name = 'suri-tm-fe'
    response = ec2_client.describe_instances()
    instance_id = None
    for reservation in response['Reservations']:
        for instance in reservation['Instances']:
            if 'Tags' in instance:
                for tag in instance['Tags']:
                    if tag['Key'] == 'Name' and tag['Value'] == instance_name and instance['State']['Name'] == 'running':
                        instance_id = instance['InstanceId']
                        break
    
    return instance_id

def create_image_ami_of_instance():
    instance_id = get_instances_running_by_name()
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    ami_name = f'suri-tm-fe-ami-{timestamp}'
    ami_response = ec2_client.create_image(
        InstanceId=instance_id,
        Name=f'AMI-{ami_name}-Timestamp-{timestamp}',
        NoReboot=True  # Set to True if you want to avoid instance reboot during the AMI creation
    )

    ami_id = ami_response['ImageId']
    return ami_id

def get_instance_info_xls_file():
    filename= 'instance_info_be.xlsx'
    dataframe = openpyxl.load_workbook(filename)
    dataframe1 = dataframe.active    
    value = dataframe1.cell(row=2, column=3).value
    return value

def create_launch_template():
    public_ip = get_instance_info_xls_file()
    print(f'The public IP address of the instance is {public_ip}')

    user_data = """#!/bin/bash
    cd /home/ubuntu/
    sudo apt install git -y
    sudo git clone https://github.com/surendergupta/TravelMemory.git
    cd /home/ubuntu/TravelMemory/frontend/
    echo "export const baseUrl = 'http://{}:3000'" > ./src/url.js
    npm install
    npm start
    """.format(public_ip)

    encoded_user_data = base64.b64encode(user_data.encode('utf-8')).decode('utf-8')

    instance_type = 't2.micro'
    ami_id = create_image_ami_of_instance()
    key_pair_name = 'suri-tm-pj'
    security_group_id = 'sg-00bc6a6d3f2de3870'
    try:
        ec2_client.create_launch_template(
            LaunchTemplateName=launch_template_name,
            VersionDescription='Initial version',
            LaunchTemplateData={
                'InstanceType': instance_type,
                'ImageId': ami_id,
                'KeyName': key_pair_name,
                'SecurityGroupIds': [security_group_id],
                'UserData': encoded_user_data
            }            
        )

        print(f'Launch template created successfully.')
    except NoCredentialsError:
        print('Credentials not available.')

# def get_target_group_arn_from_xlsx_file():
#     filename= 'elb_info.xlsx'
#     dataframe = openpyxl.load_workbook(filename)
#     dataframe1 = dataframe.active    
#     value = dataframe1.cell(row=2, column=2).value
#     return value

def check_target_group_exist():
    try:
        target_group_arn = 'suri-tm-tg'
        response = elbv2_client.describe_target_groups(Names=[target_group_arn])
        target_group = response['TargetGroups'][0]
        print(f"Target group found with ARN: {target_group['TargetGroupArn']}")        
        target_group_arn1 = target_group['TargetGroupArn']
        return target_group_arn1
    except elbv2_client.exceptions.TargetGroupNotFoundException:
        print(f"Target group not found: {target_group_arn}")

def create_asg():
    try:
        availability_zones = ['us-east-1a', 'us-east-1b', 'us-east-1c']
        target_group_arn1 = check_target_group_exist()
        asg_client.create_auto_scaling_group(
            AutoScalingGroupName=asg_name,
            LaunchTemplate={
                'LaunchTemplateName': launch_template_name,
                'Version': '$Latest',  # Use the latest version of the launch template
            },
            MinSize=min_size,
            MaxSize=max_size,
            DesiredCapacity=desired_capacity,
            AvailabilityZones=availability_zones,
            Tags=[{'Key': 'Name', 'Value': asg_name}],
            HealthCheckType='ELB',
            HealthCheckGracePeriod=300,  # seconds
            TargetGroupARNs=[target_group_arn1]
        )

        print(f'Auto Scaling Group "{asg_name}" created successfully.')
    except NoCredentialsError:
        print('Credentials not available.')

def configure_scaling_policies():
    try:
        # Configure scaling policies based on CPU utilization
        asg_client.put_scaling_policy(
            AutoScalingGroupName=asg_name,
            PolicyName='ScalePolicy',
            PolicyType='TargetTrackingScaling',
            TargetTrackingConfiguration={
                'PredefinedMetricSpecification': {
                    'PredefinedMetricType': 'ASGAverageCPUUtilization',
                },
                'TargetValue': cpu_utilization_threshold,
            }
        )
        print('Policy created successfuly in asg.')
    except NoCredentialsError:
        print('Policy not creating in asg.')

# Create Launch Configuration
create_launch_template()

# Create Auto Scaling Group
create_asg()

# Configure Scaling Policies
configure_scaling_policies()

subprocess.run(['python', 'boto3_lamfun.py'])