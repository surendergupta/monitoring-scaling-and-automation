import boto3
import base64
import openpyxl
import subprocess

aws_region = 'us-east-1'
ec2_client = boto3.resource('ec2', region_name=aws_region)

def createEC2FrontendInstance(ec2_client):
    public_ip = get_instance_info_xls_file()
    print(f'The public IP address of the instance is {public_ip}')

    instance_type = 't2.micro'
    ami_id = 'ami-0c7217cdde317cfec'
    key_pair_name = 'suri-tm-pj'
    security_group_name = 'launch-wizard-1'
    user_data_script = """#!/bin/bash
    sudo apt-get update -y
    sudo apt-get install -y ca-certificates curl gnupg
    sudo mkdir -p /etc/apt/keyrings
    curl -fsSL https://deb.nodesource.com/gpgkey/nodesource-repo.gpg.key | sudo gpg --dearmor -o /etc/apt/keyrings/nodesource.gpg
    NODE_MAJOR=18
    echo "deb [signed-by=/etc/apt/keyrings/nodesource.gpg] https://deb.nodesource.com/node_$NODE_MAJOR.x nodistro main" | sudo tee /etc/apt/sources.list.d/nodesource.list
    sudo apt-get update
    sudo apt install nodejs -y
    cd /home/ubuntu/
    sudo apt install git -y
    sudo git clone https://github.com/surendergupta/TravelMemory.git
    cd /home/ubuntu/TravelMemory/frontend/
    echo "export const baseUrl = 'http://{}:3000'" > ./src/url.js
    npm install
    npm start
    """.format(public_ip)

    user_data_encoded = base64.b64encode(user_data_script.encode()).decode('utf-8')

    response = ec2_client.create_instances(
        ImageId=ami_id,
        InstanceType=instance_type,
        KeyName=key_pair_name,
        SecurityGroups=[security_group_name],
        UserData=user_data_encoded,
        MinCount=1,
        MaxCount=1,
        TagSpecifications=[
            {
                'ResourceType': 'instance',
                'Tags': [
                    {
                        'Key': 'Name',
                        'Value': 'suri-tm-fe'
                    }
                ]
            }
        ]
    )
    return response

def terminateEC2Instance(ec2_client, instance_id):
    ec2_client.terminate_instances(InstanceIds=[instance_id])
    print("Instance terminated")

def stopEC2Instance(ec2_client, instance_id):
    ec2_client.stop_instances(InstanceIds=[instance_id])
    print("Instance stopped")


def get_instance_public_ip(instance):
    if 'PublicIpAddress' in instance:
        return instance['PublicIpAddress']
    elif 'NetworkInterfaces' in instance and instance['NetworkInterfaces']:
        # If the instance has a network interface, check for public IP in the first interface
        network_interface = instance['NetworkInterfaces'][0]
        if 'Association' in network_interface and 'PublicIp' in network_interface['Association']:
            return network_interface['Association']['PublicIp']
    return None

def get_instance_info_xls_file():
    filename= 'instance_info_be.xlsx'
    dataframe = openpyxl.load_workbook(filename)
    dataframe1 = dataframe.active    
    value = dataframe1.cell(row=2, column=3).value
    return value



respFront = createEC2FrontendInstance(ec2_client)

instance = respFront[0]
print(f"Launched frontend instance with ID: {instance.instance_id}")
instance.wait_until_running()

try:
    workbook = openpyxl.load_workbook("instance_info_fe.xlsx")
    worksheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Instance Info"
    worksheet['A1'] = "Instance ID"
    worksheet['B1'] = "Instance Type"
    worksheet['C1'] = "Public IPv4 address"
    worksheet['D1'] = "Private IP"
    worksheet['E1'] = "Launch time"
    worksheet['F1'] = "Security group name"

next_row = worksheet.max_row + 1
worksheet.cell(row=next_row, column=1, value=instance.instance_id)
worksheet.cell(row=next_row, column=2, value=instance.instance_type)
worksheet.cell(row=next_row, column=3, value=instance.public_ip_address)
worksheet.cell(row=next_row, column=4, value=instance.private_ip_address)
worksheet.cell(row=next_row, column=5, value=str(instance.launch_time))
worksheet.cell(row=next_row, column=6, value=instance.security_groups[0]['GroupName'] if instance.security_groups else '')

# Save the Excel file
workbook.save("instance_info_fe.xlsx")

# instance = respFront[0]
# instance.wait_until_running()

subprocess.run(['python', 'boto3_elb.py'])
