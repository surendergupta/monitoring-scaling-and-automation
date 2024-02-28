import boto3
import openpyxl
from botocore.exceptions import NoCredentialsError
import subprocess

aws_region = 'us-east-1'
ec2_client = boto3.client('ec2', region_name=aws_region)
elbv2_client = boto3.client('elbv2', region_name=aws_region)
vpc_id = 'vpc-08bfb22193349823a'
alb_name = 'suri-tm-lb'
alb_subnets = ['subnet-01669f7781e6f4aaf', 'subnet-005cbfd80b6b905a8', 'subnet-09be1ae1b5653e3c9', 'subnet-0e3b372e2a1558091', 'subnet-090ddef663a93d582', 'subnet-03b9d0353954f0b4a']
alb_security_groups = ['sg-00bc6a6d3f2de3870']

target_group_name = 'suri-tm-tg'
protocol = 'HTTP'
port = 80
def create_alb():
    try:
        response = elbv2_client.create_load_balancer(
            Name=alb_name,
            Subnets=alb_subnets,
            SecurityGroups=alb_security_groups,
            Scheme='internet-facing',
            Tags=[{'Key': 'Name', 'Value': alb_name}]
        )

        alb_arn = response['LoadBalancers'][0]['LoadBalancerArn']
        print(f'ALB "{alb_name}" created successfully with ARN: {alb_arn}')
        
        elbv2_client.get_waiter('load_balancer_exists').wait(Names=[alb_name])
        print(f"Load balancer {alb_name} is active.")
        
        return alb_arn
    except NoCredentialsError:
        print('Credentials not available.')

def create_target_group():
    try:
        response = elbv2_client.create_target_group(
            Name = target_group_name,
            Protocol = protocol,
            Port = 3000,
            VpcId = vpc_id,           
            TargetType = 'instance',
            Tags = [{'Key': 'Name', 'Value': target_group_name}]
        )

        target_group_arn = response['TargetGroups'][0]['TargetGroupArn']
        print(f'Target Group "{target_group_name}" created successfully with ARN: {target_group_arn}')
        
        return target_group_arn
    except NoCredentialsError:
        print('Credentials not available.')

def register_targets_with_alb(alb_arn, target_group_arn, ec2_instance_ids):
    try:
        for instance_id in ec2_instance_ids:
            elbv2_client.register_targets(
                TargetGroupArn=target_group_arn,
                Targets=[{'Id': instance_id}]
            )
            print(f"Instance {instance_id} registered with target group.")

        print(f'Registered targets with Target Group successfully.')

        elbv2_client.create_listener(
            LoadBalancerArn=alb_arn,
            Protocol=protocol,
            Port=port,
            DefaultActions=[
                {
                    'Type': 'forward',
                    'TargetGroupArn': target_group_arn,
                },
            ],
        )

        print("ALB created, target group created, instances registered, and listener with forward action added successfully.")
    except NoCredentialsError:
        print('Credentials not available.')

def get_instances_running_by_name():
    instance_name = 'suri-tm-fe'
    response = ec2_client.describe_instances()
    instance_id = None
    for reservation in response['Reservations']:
        for instance in reservation['Instances']:
            if 'Tags' in instance:
                for tag in instance['Tags']:
                    if tag['Key'] == 'Name' and tag['Value'] == instance_name and instance['State']['Name'] == 'running':
                        instance_id = instance['InstanceId']
                        break
    
    if instance_id:
        ec2_instance_ids = [instance_id]
        print(f"Instance ID for {instance_name} is {instance_id}")
        return ec2_instance_ids
    else:
        print(f"No instance with the name {instance_name} found.")

def get_instance_id_from_xlsx_file():
    filename= 'instance_info_fe.xlsx'
    dataframe = openpyxl.load_workbook(filename)
    dataframe1 = dataframe.active    
    value = dataframe1.cell(row=2, column=1).value
    return value

# Frontend Instance ID
ec2_instance_ids = get_instances_running_by_name()

# Create ALB
alb_arn = create_alb()

# Create Target Group
target_group_arn = create_target_group()

# Register EC2 instance(s) with ALB
register_targets_with_alb(alb_arn, target_group_arn, ec2_instance_ids)

try:
    workbook = openpyxl.load_workbook("elb_info.xlsx")
    worksheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "ELB Info"
    worksheet['A1'] = "ALB ARN"
    worksheet['B1'] = "Target Group ARN"

# Add instance information to the worksheet
next_row = worksheet.max_row + 1
worksheet.cell(row=next_row, column=1, value=alb_arn)
worksheet.cell(row=next_row, column=2, value=target_group_arn)

# Save the Excel file
workbook.save("elb_info.xlsx")

subprocess.run(['python', 'boto3_asg.py'])