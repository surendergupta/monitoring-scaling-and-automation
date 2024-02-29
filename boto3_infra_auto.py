import boto3
import openpyxl
import base64
import gzip
import re
import json
import time
from datetime import datetime
from botocore.exceptions import NoCredentialsError

aws_region = 'us-east-1'
ec2_client = boto3.client('ec2', region_name=aws_region)
elbv2_client = boto3.client('elbv2', region_name=aws_region)
asg_client = boto3.client('autoscaling', region_name=aws_region)
s3_client = boto3.client('s3')
iam_client = boto3.client('iam')
sns_client = boto3.client('sns', region_name=aws_region)
lambda_client = boto3.client('lambda', region_name=aws_region)

# Vpc id Its change when you are setup this file in your 
vpc_id = 'vpc-08bfb22193349823a'

instance_type = 't2.micro'
key_pair_name = 'suri-tm-pj'
security_group_id = 'sg-05c5ea9e91cc7016f'

# ELB Globle Variable
alb_name = 'suri-tm-lb'
alb_subnets = ['subnet-01669f7781e6f4aaf', 'subnet-005cbfd80b6b905a8', 'subnet-09be1ae1b5653e3c9', 'subnet-0e3b372e2a1558091', 'subnet-090ddef663a93d582', 'subnet-03b9d0353954f0b4a']
alb_security_groups = [security_group_id]
target_group_name = 'suri-tm-tg'
protocol = 'HTTP'
port = 80

# ASG Globle Variable
availability_zones = ['us-east-1a', 'us-east-1b', 'us-east-1c', 'us-east-1d', 'us-east-1e', 'us-east-1f']
launch_template_name = 'suri-tm-fe-launch-template'
asg_name = 'suri-tm-asg'
min_size = 1
max_size = 3
desired_capacity = 1

# Scaling policies configuration
cpu_utilization_threshold = 50  # Scale out if CPU utilization is above this threshold
network_traffic_threshold = 5000000  # Scale out if network traffic is above this threshold
scale_out_cooldown = 300  # seconds
scale_in_cooldown = 300  # seconds
target_group_arn1 = None

# lambda function
lambda_function_name = 'lambda-function-boto3'
snapshot_description = 'WebAppFailureSnapshot'
topic_name = 'AdminNotifications'

s3_bucket_name = 'lambda-functions-bucket-s3'

# Step 1: Set Up MongoDB
mongo_url = 'mongodb+srv://surendergupta:ceOC9UhUh5trGjEE@taskmanagementcluster.tsyynim.mongodb.net/travelMemory'


def createEC2BackendInstance():
    ami_id = 'ami-0c7217cdde317cfec'
    
    user_data_script = """#!/bin/bash
    sudo apt-get update -y
    sudo apt-get install -y ca-certificates curl gnupg
    sudo mkdir -p /etc/apt/keyrings
    curl -fsSL https://deb.nodesource.com/gpgkey/nodesource-repo.gpg.key | sudo gpg --dearmor -o /etc/apt/keyrings/nodesource.gpg
    NODE_MAJOR=18
    echo "deb [signed-by=/etc/apt/keyrings/nodesource.gpg] https://deb.nodesource.com/node_$NODE_MAJOR.x nodistro main" | sudo tee /etc/apt/sources.list.d/nodesource.list
    sudo apt-get update
    sudo apt install nodejs -y
    cd /home/ubuntu/
    sudo apt install git -y
    sudo git clone https://github.com/surendergupta/TravelMemory.git
    cd /home/ubuntu/TravelMemory/backend/
    echo "MONGO_URI='{}'" > .env
    echo "PORT=5000" >> .env
    npm install
    node index.js
    """.format(mongo_url)
    user_data_encoded = base64.b64encode(user_data_script.encode()).decode('utf-8')

    response = ec2_client.run_instances(
        ImageId=ami_id,
        InstanceType=instance_type,
        KeyName=key_pair_name,
        SecurityGroupIds=[security_group_id],
        UserData=user_data_encoded,
        MinCount=1,
        MaxCount=1,
        TagSpecifications=[
            {
                'ResourceType': 'instance',
                'Tags': [
                    {
                        'Key': 'Name',
                        'Value': 'suri-tm-be'
                    }
                ]
            }
        ]
    )
    return response

def createEC2FrontendInstance():
    public_ip = get_instance_info_xls_file()
    print(f'The public IP address of the instance is {public_ip}')

    ami_id = 'ami-0c7217cdde317cfec'
    user_data_script = """#!/bin/bash
    sudo apt-get update -y
    sudo apt-get install -y ca-certificates curl gnupg
    sudo mkdir -p /etc/apt/keyrings
    curl -fsSL https://deb.nodesource.com/gpgkey/nodesource-repo.gpg.key | sudo gpg --dearmor -o /etc/apt/keyrings/nodesource.gpg
    NODE_MAJOR=18
    echo "deb [signed-by=/etc/apt/keyrings/nodesource.gpg] https://deb.nodesource.com/node_$NODE_MAJOR.x nodistro main" | sudo tee /etc/apt/sources.list.d/nodesource.list
    sudo apt-get update
    sudo apt install nodejs -y
    cd /home/ubuntu/
    sudo apt install git -y
    sudo git clone https://github.com/surendergupta/TravelMemory.git
    cd /home/ubuntu/TravelMemory/frontend/
    echo "export const baseUrl = 'http://{}:3000'" > ./src/url.js
    npm install
    npm start
    """.format(public_ip)

    user_data_encoded = base64.b64encode(user_data_script.encode()).decode('utf-8')

    response = ec2_client.run_instances(
        ImageId=ami_id,
        InstanceType=instance_type,
        KeyName=key_pair_name,
        SecurityGroupIds=[security_group_id],
        UserData=user_data_encoded,
        MinCount=1,
        MaxCount=1,
        TagSpecifications=[
            {
                'ResourceType': 'instance',
                'Tags': [
                    {
                        'Key': 'Name',
                        'Value': 'suri-tm-fe'
                    }
                ]
            }
        ]
    )
    return response

def terminateEC2Instance(instance_id):
    ec2_client.terminate_instances(InstanceIds=[instance_id])
    print("Instance terminated")

def stopEC2Instance(instance_id):
    ec2_client.stop_instances(InstanceIds=[instance_id])
    print("Instance stopped")

def get_instance_public_ip(instance):
    if 'PublicIpAddress' in instance:
        return instance['PublicIpAddress']
    elif 'NetworkInterfaces' in instance and instance['NetworkInterfaces']:
        # If the instance has a network interface, check for public IP in the first interface
        network_interface = instance['NetworkInterfaces'][0]
        if 'Association' in network_interface and 'PublicIp' in network_interface['Association']:
            return network_interface['Association']['PublicIp']
    return None

def get_instance_info_xls_file():
    filename= 'instance_info_be.xlsx'
    dataframe = openpyxl.load_workbook(filename)
    dataframe1 = dataframe.active    
    value = dataframe1.cell(row=2, column=3).value
    return value

def createBeWorksheet(file_name, instance_id):
    ec2_resource = boto3.resource('ec2', region_name=aws_region)
    instance = ec2_resource.Instance(instance_id)
    instance.wait_until_running()
    try:
        workbook = openpyxl.load_workbook(file_name)
        worksheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Instance Info"
        worksheet['A1'] = "Instance ID"
        worksheet['B1'] = "Instance Type"
        worksheet['C1'] = "Public IPv4 address"
        worksheet['D1'] = "Private IP"
        worksheet['E1'] = "Launch time"
        worksheet['F1'] = "Security group name"

    # Add instance information to the worksheet
    next_row = worksheet.max_row + 1
    worksheet.cell(row=next_row, column=1, value=instance.instance_id)
    worksheet.cell(row=next_row, column=2, value=instance.instance_type)
    worksheet.cell(row=next_row, column=3, value=instance.public_ip_address)
    worksheet.cell(row=next_row, column=4, value=instance.private_ip_address)
    worksheet.cell(row=next_row, column=5, value=str(instance.launch_time))
    worksheet.cell(row=next_row, column=6, value=instance.security_groups[0]['GroupName'] if instance.security_groups else '')

    # Save the Excel file
    workbook.save(file_name)

def create_alb():
    try:
        response = elbv2_client.create_load_balancer(
            Name=alb_name,
            Subnets=alb_subnets,
            SecurityGroups=alb_security_groups,
            Scheme='internet-facing',
            Tags=[{'Key': 'Name', 'Value': alb_name}]
        )

        alb_arn = response['LoadBalancers'][0]['LoadBalancerArn']
        print(f'ALB "{alb_name}" created successfully with ARN: {alb_arn}')
        
        elbv2_client.get_waiter('load_balancer_exists').wait(Names=[alb_name])
        print(f"Load balancer {alb_name} is active.")
        
        return alb_arn
    except NoCredentialsError:
        print('Credentials not available.')

def create_target_group():
    try:
        response = elbv2_client.create_target_group(
            Name = target_group_name,
            Protocol = protocol,
            Port = 3000,
            VpcId = vpc_id,           
            TargetType = 'instance',
            Tags = [{'Key': 'Name', 'Value': target_group_name}]
        )

        target_group_arn = response['TargetGroups'][0]['TargetGroupArn']
        print(f'Target Group "{target_group_name}" created successfully with ARN: {target_group_arn}')
        
        return target_group_arn
    except NoCredentialsError:
        print('Credentials not available.')

def register_targets_with_alb(alb_arn, target_group_arn, ec2_instance_ids):
    try:
        for instance_id in ec2_instance_ids:
            elbv2_client.register_targets(
                TargetGroupArn=target_group_arn,
                Targets=[{'Id': instance_id}]
            )
            print(f"Instance {instance_id} registered with target group.")

        print(f'Registered targets with Target Group successfully.')

        elbv2_client.create_listener(
            LoadBalancerArn=alb_arn,
            Protocol=protocol,
            Port=port,
            DefaultActions=[
                {
                    'Type': 'forward',
                    'TargetGroupArn': target_group_arn,
                },
            ],
        )

        print("ALB created, target group created, instances registered, and listener with forward action added successfully.")
    except NoCredentialsError:
        print('Credentials not available.')

def get_instances_running_by_name():    
    instance_name = 'suri-tm-fe'
    response = ec2_client.describe_instances()
    instance_id = None
    for reservation in response['Reservations']:
        for instance in reservation['Instances']:
            if 'Tags' in instance:
                for tag in instance['Tags']:
                    if tag['Key'] == 'Name' and tag['Value'] == instance_name and instance['State']['Name'] == 'running':
                        instance_id = instance['InstanceId']
                        break
    
    if instance_id:
        ec2_instance_ids = [instance_id]
        print(f"Instance ID for {instance_name} is {instance_id}")
        return ec2_instance_ids
    else:
        print(f"No instance with the name {instance_name} found.")

def get_instance_id_from_xlsx_file():
    filename= 'instance_info_fe.xlsx'
    dataframe = openpyxl.load_workbook(filename)
    dataframe1 = dataframe.active    
    value = dataframe1.cell(row=2, column=1).value
    return value

def createELBWorksheet(file_name, alb_arn, target_group_arn):
    try:
        workbook = openpyxl.load_workbook(file_name)
        worksheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "ELB Info"
        worksheet['A1'] = "ALB ARN"
        worksheet['B1'] = "Target Group ARN"

    # Add instance information to the worksheet
    next_row = worksheet.max_row + 1
    worksheet.cell(row=next_row, column=1, value=alb_arn)
    worksheet.cell(row=next_row, column=2, value=target_group_arn)

    # Save the Excel file
    workbook.save(file_name)

def create_image_ami_of_instance():
    instance_id = get_instances_running_by_name()
    # print(f"instance_id : {instance_id}")
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    ami_name = f'suri-tm-fe-ami-{timestamp}'
    
    ami_response = ec2_client.create_image(
        InstanceId=instance_id[0],
        Name=f'AMI-{ami_name}-Timestamp-{timestamp}',
        NoReboot=True  # Set to True if you want to avoid instance reboot during the AMI creation
    )

    ami_id = ami_response['ImageId']
    return ami_id

def get_instance_ip_info_xls_file(filename):
    #filename= 'instance_info_be.xlsx'
    dataframe = openpyxl.load_workbook(filename)
    dataframe1 = dataframe.active    
    value = dataframe1.cell(row=2, column=3).value
    return value

def create_launch_template():
    public_ip = get_instance_ip_info_xls_file('instance_info_be.xlsx')
    print(f'The public IP address of the instance is {public_ip}')

    user_data = """#!/bin/bash
    cd /home/ubuntu/
    sudo apt install git -y
    sudo git clone https://github.com/surendergupta/TravelMemory.git
    cd /home/ubuntu/TravelMemory/frontend/
    echo "export const baseUrl = 'http://{}:3000'" > ./src/url.js
    npm install
    npm start
    """.format(public_ip)

    encoded_user_data = base64.b64encode(user_data.encode('utf-8')).decode('utf-8')    
    ami_id = create_image_ami_of_instance()
    
    try:
        ec2_client.create_launch_template(
            LaunchTemplateName=launch_template_name,
            VersionDescription='Initial version',
            LaunchTemplateData={
                'InstanceType': instance_type,
                'ImageId': ami_id,
                'KeyName': key_pair_name,
                'SecurityGroupIds': [security_group_id],
                'UserData': encoded_user_data
            },
            TagSpecifications=[
                {
                    'ResourceType': 'launch-template',
                    'Tags': [
                        {
                            'Key': 'Name',
                            'Value': 'suri-tm-asg-fe'
                        }
                    ]
                }
            ]
        )

        print(f'Launch template created successfully.')
    except NoCredentialsError:
        print('Credentials not available.')

def check_target_group_exist():
    try:
        target_group_arn = target_group_name
        response = elbv2_client.describe_target_groups(Names=[target_group_arn])
        target_group = response['TargetGroups'][0]
        print(f"Target group found with ARN: {target_group['TargetGroupArn']}")        
        target_group_arn1 = target_group['TargetGroupArn']
        return target_group_arn1
    except elbv2_client.exceptions.TargetGroupNotFoundException:
        print(f"Target group not found: {target_group_arn}")

def create_asg():
    try:
        
        target_group_arn1 = check_target_group_exist()
        asg_client.create_auto_scaling_group(
            AutoScalingGroupName=asg_name,
            LaunchTemplate={
                'LaunchTemplateName': launch_template_name,
                'Version': '$Latest',  # Use the latest version of the launch template
            },
            MinSize=min_size,
            MaxSize=max_size,
            DesiredCapacity=desired_capacity,
            AvailabilityZones=availability_zones,
            Tags=[{'Key': 'Name', 'Value': asg_name}],
            HealthCheckType='ELB',
            HealthCheckGracePeriod=300,  # seconds
            TargetGroupARNs=[target_group_arn1]
        )

        print(f'Auto Scaling Group "{asg_name}" created successfully.')
    except NoCredentialsError:
        print('Credentials not available.')

def configure_scaling_policies():
    try:
        # Configure scaling policies based on CPU utilization
        asg_client.put_scaling_policy(
            AutoScalingGroupName=asg_name,
            PolicyName='ScalePolicy',
            PolicyType='TargetTrackingScaling',
            TargetTrackingConfiguration={
                'PredefinedMetricSpecification': {
                    'PredefinedMetricType': 'ASGAverageCPUUtilization',
                },
                'TargetValue': cpu_utilization_threshold,
            }
        )
        print('Policy created successfuly in asg.')
    except NoCredentialsError:
        print('Policy not creating in asg.')

def get_alb_arn_from_xlsx_file():
    filename= 'elb_info.xlsx'
    dataframe = openpyxl.load_workbook(filename)
    dataframe1 = dataframe.active    
    value = dataframe1.cell(row=4, column=2).value
    return value

def getTopicArn():
    response = sns_client.list_topics()
    topic_arn = None
    for topic in response['Topics']:
        # Check if the topic name matches the desired topic name        
        topic_attributes = sns_client.get_topic_attributes(TopicArn=topic['TopicArn'])
        topic_name = topic_attributes['Attributes'].get('DisplayName', topic['TopicArn'])    
        if topic_name == topic_name:
            topic_arn = topic['TopicArn']
            break
    if topic_arn is not None:
        print("ARN of the topic:", topic_arn)
        return topic_arn        
    else:
        print("Topic not found.")
        return topic_arn

def capture_instance_snapshot(instance_id):
    # Create snapshot with a descriptive name
    snapshot_description = f"Snapshot for troubleshooting instance {instance_id}"
    snapshot = ec2_client.create_snapshot(VolumeId=instance_id, Description=snapshot_description)
    snapshot_id = snapshot['SnapshotId']
    print(f"Snapshot created: {snapshot_id}")
    return snapshot_id

def send_notification(instance_id, snapshot_id):
    # GET SNS topic ARN
    topic_arn = getTopicArn()
    # Send a notification through SNS to administrators
    sns_client.publish(
        TopicArn=topic_arn,
        Subject='Web Application Health Check Alert',
        Message=f'Web application health check failed for instance {instance_id}. Instance terminated. Snapshot ID: {snapshot_id}',
    )

def lambda_handler(event, context, alb_target_group_arn):
    try:
        if event:
            for record in event['Records']:
                bucket_name = record['s3']['bucket']['name']
                object_key = record['s3']['object']['key']

                # Download the log file from S3
                response = s3_client.get_object(Bucket=bucket_name, Key=object_key)
                log_data = response['Body'].read().decode('utf-8')

                if object_key.endswith('.gz'):
                    log_data = gzip.decompress(log_data)

                    log_lines = log_data.decode('utf-8').split('\n')
                    for log_line in log_lines:
                        # Analyze log line for suspicious activities or high traffic
                        if is_health_issue(log_data):
                            subject = 'Web Application Health Issue detected Alert'
                            message = f"Health issue detected in log file: {object_key}"
                            send_notification(subject, message)
                            
                        elif is_scaling_event(log_data):
                            subject = 'Web Application Scaling detected Alert'
                            message = f"Scaling event detected in log file: {object_key}"
                            send_notification(subject, message)
                            
                        elif is_high_traffic(log_data):
                            subject = 'Web Application High Traffic detected Alert'
                            message = f'High traffic detected in ALB access log: {object_key}'
                            send_notification(subject, message)
                            
                        elif is_suspicious(log_line):
                            # Send notification via SNS
                            subject = 'Web Application DDoS attack detected Alert'
                            message = 'Web application Potential DDoS attack detected.'
                            send_notification(subject, message)
                else:
                    # Perform log analysis (example: check for high traffic)
                    if 'HighTrafficKeyword' in log_data:
                        # Send a notification via SNS
                        subject = 'High Traffic Alert'
                        message = f'High traffic detected in ALB access log: {object_key}'
                        send_notification(subject, message) 
        else:
                    
            # Describe target health for the ALB
            response = elbv2_client.describe_target_health(TargetGroupArn=alb_target_group_arn)
            
            for target in response['TargetHealthDescriptions']:
                target_id = target['Target']['Id']
                target_health = target['TargetHealth']['State']            
                
                ec2_response = ec2_client.describe_instances(InstanceIds=[target_id])
                volume_id = ec2_response['Reservations'][0]['Instances'][0]['BlockDeviceMappings'][0]['Ebs']['VolumeId']
                print("Volume ID:", volume_id)

                if target_health == 'unhealthy':
                    print(f'Instance {target_id} is unhealthy. Taking corrective actions...')
                    
                    # Capture snapshot
                    snapshot_id = capture_instance_snapshot(volume_id)

                    # Terminate instance
                    terminateEC2Instance(target_id)
                    
                    # Send notification
                    send_notification(target_id, snapshot_id)                
                    
                    print('Notification sent to administrators.')
                else:
                    print('All Instance is Healthy no need to send notification to administrators.')
    except NoCredentialsError:
        print('Catch error in lambda function.')

def configure_alb_logging(alb_arn):
    try:
        attach_s3_policy_to_role('roleS3Access')

        elbv2_client.modify_load_balancer_attributes(
            LoadBalancerArn=alb_arn,
            Attributes=[
                {
                    'Key': 'access_logs.s3.enabled',
                    'Value': 'true',
                },
                {
                    'Key': 'access_logs.s3.bucket',
                    'Value': s3_bucket_name,
                },
                {
                    'Key': 'access_logs.s3.prefix',
                    'Value': 's3-alb-access-logs-v1',
                },
            ],
        )
        print(f'ALB access logs configured successfully. Log files will be stored in S3 bucket: {s3_bucket_name}')

    except NoCredentialsError:
        print('Credentials not available.')

def attach_s3_policy_to_role(role_name):
    policy_document = {
        "Version": "2012-10-17",
        "Statement": [
            {
                "Effect": "Allow",
                "Action": "elasticloadbalancing:ModifyLoadBalancerAttributes",
                "Resource": "arn:aws:elasticloadbalancing:us-east-1:060095847722:loadbalancer/app/suri-tm-lb/a0b625daa05bd7e2"            
            },
            {
                "Effect": "Allow",
                "Action": [
                    "s3:GetBucketLogging",
                    "s3:PutBucketLogging",
                    "s3:GetBucketAcl",
                    "s3:PutBucketAcl",
                    "s3:CreateBucket",
                    "s3:ListBucket",
                    "s3:PutObject"
                ],
                "Resource": [
                    f"arn:aws:s3:::{s3_bucket_name}",
                    f"arn:aws:s3:::{s3_bucket_name}/*"
                ]
            },
            {
                "Effect": "Allow",
                "Action": [
                    "s3:GetObject",
                    "s3:PutObject"
                ],
                "Resource": [
                    f"arn:aws:s3:::{s3_bucket_name}/*"
                ]
            }
        ]
    }

    try:
        response = iam_client.put_role_policy(
            RoleName=role_name,
            PolicyName='S3AccessPolicy',
            PolicyDocument=json.dumps(policy_document)
        )
        print(response)
        print("IAM policy attached successfully.")
    except NoCredentialsError:
        print("Credentials not available.")

def is_health_issue(log_data):
    if "500 Internal Server Error" in log_data:
        return True
    
    if "CRITICAL" in log_data:
        return True
    return False

def is_scaling_event(log_data):
    if "Scaling event" in log_data:
        return True
    
    if "Increased traffic" in log_data:
        return True
    
    return False

def is_high_traffic(log_data):
    log_lines = log_data.split('\n')
    request_count = sum(1 for line in log_lines if 'GET' in line or 'POST' in line)
    traffic_threshold = 1000
    
    if request_count > traffic_threshold:
        return True
    
    ip_request_count = {}
    for line in log_lines:
        if 'GET' in line or 'POST' in line:
            ip = line.split()[0]  # Assuming IP address is the first part of the log entry
            ip_request_count[ip] = ip_request_count.get(ip, 0) + 1
    high_traffic_threshold = 100

    for ip, count in ip_request_count.items():
        if count > high_traffic_threshold:
            return True
        
    return False

def is_suspicious(log_line):
    ip_address = re.search(r'(\d+\.\d+\.\d+\.\d+)', log_line).group(1)
    count = log_line.count(ip_address)
    if count > 1000:
        return True
    else:
        return False


def get_alb_arn_from_xlsx_file():
    filename= 'elb_info.xlsx'
    dataframe = openpyxl.load_workbook(filename)
    dataframe1 = dataframe.active    
    value = dataframe1.cell(row=2, column=1).value
    return value

def get_alb_tg_arn_from_xlsx_file():
    filename= 'elb_info.xlsx'
    dataframe = openpyxl.load_workbook(filename)
    dataframe1 = dataframe.active    
    value = dataframe1.cell(row=2, column=2).value
    return value


def create_s3_bucket():
    try:
        s3_client.create_bucket(Bucket=s3_bucket_name, CreateBucketConfiguration={'LocationConstraint': aws_region})
        print(f'S3 bucket "{s3_bucket_name}" created successfully.')
    except NoCredentialsError:
        print('Credentials not available.')

def create_lambda_function():
    try:
        function_code_file_path = './lambda_function.py'

        # Read the function code from the file
        with open(function_code_file_path, 'r') as file:
            function_code = file.read()

        function_code = '''
        def lambda_handler(event, context):
            # Your Lambda function logic here
            return {
                'statusCode': 200,
                'body': 'Hello from Lambda!'
            }
        '''
        response = lambda_client.create_function(
            FunctionName=lambda_function_name,
            Runtime='python3.8',
            Role='arn:aws:iam::060095847722:role/lambda-role-boto3',
            Handler= 'lambda_handler',  # Replace with your Lambda handler function
            Code={
                'ZipFile': function_code.encode()
            },
            Timeout=30,
            MemorySize=128,
            Environment={'Variables': {'AWSREGION': aws_region}},
            Tags={'Key': 'Name', 'Value': lambda_function_name}
        )

        print(f'Lambda function "{lambda_function_name}" created successfully.')
        return response['FunctionArn']
    except NoCredentialsError:
        print('Credentials not available.')

def create_sns_topic(topic_name):
    # Create SNS topic
    try:
        response = sns_client.create_topic(Name=topic_name)
        topic_arn = response['TopicArn']
        print(f'SNS topic "{topic_name}" created successfully with ARN: {topic_arn}')

        sns_client.set_topic_attributes(
            TopicArn=topic_arn,
            AttributeName='DisplayName',
            AttributeValue='Admin Notifications'
        )

        return topic_arn
    except NoCredentialsError:
        print('Credentials not available.')

def send_notification(topic_arn, subject, message):
    # Send a notification through SNS to administrators
    sns_client.publish(
        TopicArn=topic_arn,
        Subject=subject,
        Message=message,
    )
    print("Notification sent.")


def subscribe_lambda_to_topic(topic_arn, lambda_function_arn, protocol='lambda'):
    try:
        response = sns_client.subscribe(
            TopicArn=topic_arn,
            Protocol=protocol,
            Endpoint=lambda_function_arn
        )

        print(f'Lambda function subscribed to SNS topic "{topic_arn}" successfully.')
    except NoCredentialsError:
        print('Credentials not available.')

def delete_resources():
    try:
        # Delete Auto Scaling Group
        asg_client.delete_auto_scaling_group(AutoScalingGroupName=asg_name, ForceDelete=True)

        # Delete Launch Configuration
        asg_client.delete_launch_configuration(LaunchConfigurationName=launch_template_name)

        # Deregister and delete targets from the ALB Target Group
        target_group_arn = elbv2_client.describe_target_groups(Names=[asg_name])['TargetGroups'][0]['TargetGroupArn']
        targets = elbv2_client.describe_target_health(TargetGroupArn=target_group_arn)['TargetHealthDescriptions']
        target_ids = [target['Target']['Id'] for target in targets]
        if target_ids:
            elbv2_client.deregister_targets(TargetGroupArn=target_group_arn, Targets=[{'Id': target_id} for target_id in target_ids])

        # Delete the ALB
        elbv2_client.delete_load_balancer(LoadBalancerArn=alb_name)

        # Delete Lambda function
        lambda_client.delete_function(FunctionName=lambda_function_name)

        # Delete S3 bucket
        s3_client.delete_bucket(Bucket=s3_bucket_name)

        print('Infrastructure deleted successfully.')
    except NoCredentialsError:
        print('Credentials not available.')

if __name__ == '__main__':
    # Create Backend Instance
    # beResponse = createEC2BackendInstance()
    # be_instance_id = beResponse['Instances'][0]['InstanceId']
    # print(f"Launched backend instance with ID: {be_instance_id}")

    # Save Backend Instance Information in worksheet
    # createBeWorksheet('instance_info_be.xlsx', be_instance_id)
    
    # Create Fontend Instance
    # feResponse = createEC2FrontendInstance()
    # fe_instance_id = feResponse['Instances'][0]['InstanceId']
    # print(f"Launched frontend instance with ID: {fe_instance_id}")

    # # Save frontend Instance Information in worksheet
    # createBeWorksheet('instance_info_fe.xlsx', fe_instance_id)    

    # # Frontend Instance ID
    # ec2_instance_ids = get_instances_running_by_name()

    # # Create ALB
    # alb_arn = create_alb()

    # # Create Target Group
    # target_group_arn = create_target_group()

    # # Register EC2 instance(s) with ALB
    # register_targets_with_alb(alb_arn, target_group_arn, ec2_instance_ids)

    # # Save ELB Information in worksheet
    # createELBWorksheet('elb_info.xlsx', alb_arn, target_group_arn)

    # # Create Launch Configuration
    # create_launch_template()

    # # Create Auto Scaling Group
    # create_asg()

    # # Configure Scaling Policies
    # configure_scaling_policies()
    
     # # Configure ALB logging
    # alb_arn = get_alb_arn_from_xlsx_file()
    # print(alb_arn)
    # configure_alb_logging(alb_arn)    

    lambda_function_arn = create_lambda_function()
    
    # time.sleep(60)

    # admin_notification_topic_arn = create_sns_topic('AdminNotifications')
    # health_issues_topic_arn = create_sns_topic('health-issues')
    # scaling_events_topic_arn = create_sns_topic('scaling-events')
    # high_traffic_topic_arn = create_sns_topic('high-traffic')

    # Subscribe Lambda function to relevant SNS topics
    # subscribe_lambda_to_topic(admin_notification_topic_arn, lambda_function_arn)
    # subscribe_lambda_to_topic(health_issues_topic_arn, lambda_function_arn)
    # subscribe_lambda_to_topic(scaling_events_topic_arn, lambda_function_arn)
    # subscribe_lambda_to_topic(high_traffic_topic_arn, lambda_function_arn)

   
    #botocore.errorfactory.InvalidConfigurationRequestException: An error occurred (InvalidConfigurationRequest) when calling the ModifyLoadBalancerAttributes operation: Access Denied for bucket: s3-alb-access-logs-v1. Please check S3bucket permission
    # sample_event = {
    #     'Records': [
    #         {
    #             's3': {
    #                 'bucket': {
    #                     'name': s3_bucket_name,
    #                 },
    #                 'object': {
    #                     'key': 'sample-alb-access-log.log',
    #                 },
    #             },
    #         },
    #     ],
    # }
    
    # lambda_handler(sample_event, {})

    # # get target group ARN
    # alb_target_group_arn = get_alb_tg_arn_from_xlsx_file()
    # lambda_handler({}, {}, alb_target_group_arn)

    # delete_resources()
    delete_resources()

    